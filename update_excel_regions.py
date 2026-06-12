"""
update_excel_regions.py
Adds 6 new region sections to Voice_Queue_Intraday.xlsx between the
'-- Client ProSupport EMEA --' separator row and the COLOUR LEGEND block.

Handles openpyxl's known issue where insert_rows() does not update
merged-cell ranges automatically.
"""

import sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

FILE_PATH = (
    r"D:\OneDrive - Aligned Automation Services Private Limited"
    r"\Desktop\RTA_V3\Voice_Queue_Intraday.xlsx"
)

NEW_DATE = "2026-06-10"

NEW_REGIONS = [
    {
        "name": "Client ProSupport HKG",
        "rows": [
            ("Client ProSupport HKG", NEW_DATE, "08:00", "88.3%", 15, 14, "0.9%", 13, "101.2%", "0:23:15", "72.1%", "88.5%", "27.4%", "0:08:42", "0:02:11", "27.4%", "0.6%", "2.1%", "1.2%", "0.5%", "0.7%", "4.8%", "0.5%", "0.3%", "0.2%"),
            ("Client ProSupport HKG", NEW_DATE, "08:30", "86.5%", 17, 15, "1.1%", 15, "98.7%", "0:23:44", "81.4%", "90.2%", "25.8%", "0:07:15", "0:01:58", "25.8%", "0.4%", "1.8%", "2.0%", "0.3%", "0.5%", "5.1%", "0.4%", "0.2%", "0.1%"),
            ("Client ProSupport HKG", NEW_DATE, "09:00", "90.7%", 13, 12, "2.2%", 12, "97.5%", "0:23:02", "78.6%", "91.7%", "30.1%", "0:05:33", "0:01:44", "30.1%", "0.3%", "3.2%", "0.9%", "1.1%", "0.4%", "4.6%", "0.3%", "0.1%", "0.2%"),
        ],
    },
    {
        "name": "Client ProSupport MYS",
        "rows": [
            ("Client ProSupport MYS", NEW_DATE, "08:00", "87.2%", 14, 13, "0.8%", 12, "100.3%", "0:22:48", "70.5%", "87.9%", "26.3%", "0:09:05", "0:02:22", "26.3%", "0.5%", "2.3%", "1.4%", "0.6%", "0.6%", "4.5%", "0.6%", "0.4%", "0.1%"),
            ("Client ProSupport MYS", NEW_DATE, "08:30", "84.9%", 16, 14, "1.3%", 14, "99.1%", "0:22:31", "83.2%", "89.4%", "24.7%", "0:10:18", "0:02:05", "24.7%", "0.3%", "1.9%", "2.3%", "0.4%", "0.4%", "4.9%", "0.5%", "0.2%", "0.0%"),
            ("Client ProSupport MYS", NEW_DATE, "09:00", "89.1%", 12, 11, "1.9%", 11, "96.8%", "0:22:59", "76.3%", "90.8%", "28.5%", "0:06:47", "0:01:51", "28.5%", "0.2%", "3.5%", "1.1%", "1.3%", "0.5%", "4.3%", "0.4%", "0.1%", "0.3%"),
        ],
    },
    {
        "name": "Client ProSupport KOR",
        "rows": [
            ("Client ProSupport KOR", NEW_DATE, "08:00", "89.4%", 17, 16, "1.0%", 15, "102.1%", "0:23:28", "74.8%", "89.3%", "28.2%", "0:07:54", "0:02:08", "28.2%", "0.7%", "2.0%", "1.3%", "0.5%", "0.8%", "5.0%", "0.6%", "0.3%", "0.2%"),
            ("Client ProSupport KOR", NEW_DATE, "08:30", "87.6%", 19, 17, "0.7%", 17, "97.8%", "0:23:51", "86.1%", "91.0%", "26.9%", "0:06:21", "0:01:47", "26.9%", "0.4%", "1.7%", "2.5%", "0.3%", "0.6%", "5.4%", "0.5%", "0.2%", "0.1%"),
            ("Client ProSupport KOR", NEW_DATE, "09:00", "91.8%", 15, 14, "2.5%", 14, "98.4%", "0:23:09", "80.4%", "92.5%", "31.7%", "0:04:59", "0:01:55", "31.7%", "0.2%", "3.4%", "0.8%", "1.4%", "0.5%", "4.7%", "0.3%", "0.0%", "0.3%"),
        ],
    },
    {
        "name": "Client ProSupport THA",
        "rows": [
            ("Client ProSupport THA", NEW_DATE, "08:00", "85.8%", 13, 12, "0.6%", 11, "99.5%", "0:21:37", "68.9%", "86.4%", "25.1%", "0:10:31", "0:02:44", "25.1%", "0.5%", "2.4%", "1.6%", "0.7%", "0.7%", "4.2%", "0.7%", "0.5%", "0.2%"),
            ("Client ProSupport THA", NEW_DATE, "08:30", "83.4%", 15, 13, "1.5%", 13, "97.2%", "0:21:14", "79.7%", "88.6%", "23.4%", "0:11:44", "0:02:19", "23.4%", "0.3%", "2.0%", "2.7%", "0.5%", "0.4%", "4.6%", "0.6%", "0.3%", "0.0%"),
            ("Client ProSupport THA", NEW_DATE, "09:00", "88.2%", 11, 10, "1.7%", 10, "95.9%", "0:21:52", "74.2%", "89.9%", "27.8%", "0:07:22", "0:02:01", "27.8%", "0.2%", "3.7%", "1.2%", "1.6%", "0.6%", "4.0%", "0.4%", "0.1%", "0.4%"),
        ],
    },
    {
        "name": "Client ProSupport BRA",
        "rows": [
            ("Client ProSupport BRA", NEW_DATE, "08:00", "87.9%", 19, 18, "0.8%", 17, "103.4%", "0:23:41", "73.5%", "88.1%", "27.6%", "0:08:17", "0:02:15", "27.6%", "0.6%", "2.2%", "1.5%", "0.5%", "0.7%", "4.9%", "0.6%", "0.4%", "0.1%"),
            ("Client ProSupport BRA", NEW_DATE, "08:30", "86.1%", 21, 19, "1.2%", 18, "100.6%", "0:23:18", "84.3%", "90.5%", "25.5%", "0:07:43", "0:01:52", "25.5%", "0.4%", "1.8%", "2.4%", "0.4%", "0.5%", "5.3%", "0.5%", "0.2%", "0.1%"),
            ("Client ProSupport BRA", NEW_DATE, "09:00", "90.3%", 17, 16, "2.1%", 16, "99.2%", "0:23:55", "79.8%", "92.1%", "32.4%", "0:05:08", "0:01:48", "32.4%", "0.3%", "3.6%", "0.7%", "1.2%", "0.5%", "4.8%", "0.3%", "0.1%", "0.3%"),
        ],
    },
    {
        "name": "Client ProSupport TWN",
        "rows": [
            ("Client ProSupport TWN", NEW_DATE, "08:00", "88.7%", 15, 14, "1.1%", 13, "100.8%", "0:22:22", "71.9%", "88.8%", "26.9%", "0:08:55", "0:02:18", "26.9%", "0.5%", "2.1%", "1.3%", "0.6%", "0.7%", "4.7%", "0.6%", "0.3%", "0.1%"),
            ("Client ProSupport TWN", NEW_DATE, "08:30", "86.4%", 17, 15, "0.9%", 15, "98.3%", "0:22:05", "82.7%", "90.9%", "25.2%", "0:09:28", "0:02:03", "25.2%", "0.4%", "1.9%", "2.2%", "0.3%", "0.5%", "5.0%", "0.5%", "0.2%", "0.0%"),
            ("Client ProSupport TWN", NEW_DATE, "09:00", "90.5%", 13, 12, "2.0%", 12, "97.1%", "0:22:47", "77.4%", "91.3%", "29.7%", "0:06:14", "0:01:56", "29.7%", "0.2%", "3.3%", "1.0%", "1.4%", "0.5%", "4.5%", "0.3%", "0.1%", "0.2%"),
        ],
    },
]

NEW_NOTE_SUFFIX = (
    " | HK→'Client ProSupport HKG'"
    " | MY→'Client ProSupport MYS'"
    " | KR→'Client ProSupport KOR'"
    " | TH→'Client ProSupport THA'"
    " | BR→'Client ProSupport BRA'"
    " | TW→'Client ProSupport TWN'"
)


def parse_range(rng_str):
    """Parse a merged cell range string like 'A7:Y7' into (min_col, min_row, max_col, max_row)."""
    left, right = rng_str.split(":")
    # left like 'A7', right like 'Y7'
    import re
    m = re.match(r"([A-Z]+)(\d+)", left)
    min_col = column_index_from_string(m.group(1))
    min_row = int(m.group(2))
    m = re.match(r"([A-Z]+)(\d+)", right)
    max_col = column_index_from_string(m.group(1))
    max_row = int(m.group(2))
    return min_col, min_row, max_col, max_row


def main():
    print(f"Loading workbook: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH)

    if "Daily" not in wb.sheetnames:
        print("ERROR: Sheet 'Daily' not found. Aborting.")
        sys.exit(1)

    ws = wb["Daily"]
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"Sheet 'Daily': {max_row} rows x {max_col} columns")

    # ── Find anchor rows ──────────────────────────────────────────────────────
    emea_sep_row = None
    colour_legend_row = None
    agent3_row = None

    for row_idx in range(1, max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()
        if "── Client ProSupport EMEA ──" in cell_str or cell_str == "── Client ProSupport EMEA ──":
            emea_sep_row = row_idx
            print(f"  Found EMEA separator at row {row_idx}")
        if cell_str == "COLOUR LEGEND" and emea_sep_row is not None and colour_legend_row is None:
            colour_legend_row = row_idx
            print(f"  Found COLOUR LEGEND at row {row_idx}")
        if "AGENT 3 USAGE" in cell_str:
            agent3_row = row_idx
            print(f"  Found AGENT 3 USAGE note at row {row_idx}")

    # Also try ASCII fallback detection
    if emea_sep_row is None:
        for row_idx in range(1, max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val is None:
                continue
            cell_str = str(cell_val)
            if "EMEA" in cell_str and "Client ProSupport" in cell_str and ws.cell(row=row_idx, column=2).value is None:
                emea_sep_row = row_idx
                print(f"  Found EMEA separator (fallback) at row {row_idx}: {cell_str!r}")
                break

    if emea_sep_row is None:
        print("ERROR: Could not find EMEA separator row. Aborting.")
        sys.exit(1)
    if colour_legend_row is None:
        print("ERROR: Could not find 'COLOUR LEGEND' row. Aborting.")
        sys.exit(1)

    # ── Snapshot existing merged-cell ranges before insertion ────────────────
    # openpyxl insert_rows does NOT reliably update merged cell references for
    # rows >= insert_at. We snapshot them, remove all, insert rows, then
    # re-add with corrected row numbers.
    insert_at = emea_sep_row + 1
    num_new_rows = len(NEW_REGIONS) * 4  # 3 data + 1 separator per region

    print(f"\nMerged cells before insertion:")
    merged_snapshot = []  # list of (min_col, orig_min_row, max_col, orig_max_row)
    for rng in list(ws.merged_cells.ranges):
        rng_str = str(rng)
        min_col, min_row, max_col, max_row = parse_range(rng_str)
        merged_snapshot.append((min_col, min_row, max_col, max_row))
        print(f"  {rng_str}  (rows {min_row}-{max_row})")

    # Remove all existing merges
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    print(f"  Removed {len(merged_snapshot)} merged ranges")

    # ── Insert blank rows ────────────────────────────────────────────────────
    print(f"\nInserting {num_new_rows} rows at row {insert_at}")
    ws.insert_rows(insert_at, amount=num_new_rows)

    # ── Re-apply merged cells with corrected row numbers ─────────────────────
    print("Re-applying merged cells with shifted row numbers:")
    for (min_col, min_row, max_col, max_row) in merged_snapshot:
        # Shift rows that are at or after insert_at
        new_min_row = min_row + num_new_rows if min_row >= insert_at else min_row
        new_max_row = max_row + num_new_rows if max_row >= insert_at else max_row
        min_col_letter = get_column_letter(min_col)
        max_col_letter = get_column_letter(max_col)
        new_range = f"{min_col_letter}{new_min_row}:{max_col_letter}{new_max_row}"
        ws.merge_cells(new_range)
        print(f"  Re-merged: {min_col_letter}{min_row}:{max_col_letter}{max_row} -> {new_range}")

    # ── Write data into inserted rows ────────────────────────────────────────
    rows_to_write = []
    for region in NEW_REGIONS:
        for data_row in region["rows"]:
            rows_to_write.append(list(data_row))
        sep_text = f"── {region['name']} ──"
        sep_row = [sep_text] + [None] * (25 - 1)
        rows_to_write.append(sep_row)

    print(f"\nWriting {len(rows_to_write)} data rows into inserted positions:")
    for offset, row_data in enumerate(rows_to_write):
        target_row = insert_at + offset
        for col_idx, value in enumerate(row_data, start=1):
            try:
                ws.cell(row=target_row, column=col_idx, value=value)
            except AttributeError:
                # Should not happen now that we cleared merges, but log if it does
                print(f"  WARNING: Could not write to row {target_row}, col {col_idx} (merged cell?)")

    print(f"  Done writing rows {insert_at} to {insert_at + len(rows_to_write) - 1}")

    # ── Update AGENT 3 USAGE note ─────────────────────────────────────────────
    new_agent3_row = agent3_row + num_new_rows if agent3_row else None
    if new_agent3_row:
        old_val = ws.cell(row=new_agent3_row, column=1).value or ""
        new_val = str(old_val).rstrip() + NEW_NOTE_SUFFIX
        ws.cell(row=new_agent3_row, column=1, value=new_val)
        safe_val = new_val.encode("ascii", errors="replace").decode("ascii")
        print(f"\nUpdated AGENT 3 USAGE note at row {new_agent3_row}:")
        print(f"  {safe_val}")
    else:
        print("\nWARNING: AGENT 3 USAGE row not found.")

    # ── Save ──────────────────────────────────────────────────────────────────
    print(f"\nSaving to: {FILE_PATH}")
    wb.save(FILE_PATH)
    print("Saved successfully.")

    # ── Verification ─────────────────────────────────────────────────────────
    print("\n-- Verification --")
    wb2 = openpyxl.load_workbook(FILE_PATH)
    ws2 = wb2["Daily"]
    print(f"  Total rows: {ws2.max_row}")

    print("\n  New region rows (20-43):")
    for r in range(20, 44):
        v1 = ws2.cell(row=r, column=1).value
        v2 = ws2.cell(row=r, column=2).value  # Row_date
        v3 = ws2.cell(row=r, column=3).value  # Starting
        v4 = ws2.cell(row=r, column=4).value  # SL
        v5 = ws2.cell(row=r, column=5).value  # callsoffered
        v1_safe = str(v1).encode("ascii", errors="replace").decode("ascii") if v1 else "None"
        print(f"  Row {r:3d}: {v1_safe} | date={v2} | start={v3} | SL={v4} | calls={v5}")

    print("\n  Merged cells in saved file:")
    for mc in ws2.merged_cells.ranges:
        print(f"    {mc}")

    note_val = ws2.cell(row=ws2.max_row, column=1).value or ""
    safe_note = note_val.encode("ascii", errors="replace").decode("ascii")
    print(f"\n  AGENT 3 note (row {ws2.max_row}):")
    print(f"  {safe_note}")


if __name__ == "__main__":
    main()
