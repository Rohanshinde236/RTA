import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
FILE_PATH = r'D:\OneDrive - Aligned Automation Services Private Limited\Desktop\RTA_V3\Voice_Queue_Intraday.xlsx'
wb = openpyxl.load_workbook(FILE_PATH)
ws = wb['Daily']
print(f'Total rows: {ws.max_row}')
print()
print('All rows (col1 | col2 | col3 | col4 | col5):')
for r in range(1, ws.max_row+1):
    vals = [ws.cell(row=r, column=c).value for c in range(1,6)]
    safe = [str(v).encode('ascii','replace').decode('ascii') if v else 'None' for v in vals]
    sep = " | "
    print(f'  {r:3d}: {sep.join(safe)}')

print()
print('Merged cells:')
for mc in sorted(ws.merged_cells.ranges, key=lambda x: str(x)):
    print(f'  {mc}')

print()
print('AGENT 3 note (last row):')
note = ws.cell(row=ws.max_row, column=1).value or ""
safe_note = note.encode('ascii','replace').decode('ascii')
print(f'  {safe_note}')
